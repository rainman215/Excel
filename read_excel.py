
# from openpyxl import load_workbook
import openpyxl
wb = openpyxl.load_workbook('123.xlsx')
# print wb.sheetnames
# sheet = wb.active
# b4 = sheet.cell(row=2, column=1)
# print b4.value
sheet = wb['Sheet2']
print sheet["A2"].value